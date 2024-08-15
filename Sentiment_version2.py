import tweepy
from pyspark import SparkContext
from pyspark.sql import SparkSession
from pyspark.streaming import StreamingContext  # Import the StreamingContext package
from textblob import TextBlob
import tkinter as tk
import openpyxl

# Initialize SparkContext
sc = SparkContext("local[*]", "TwitterSentimentAnalysis")

# Create SparkSession
spark = SparkSession(sc)

consumer_key = ''
consumer_secret = ''
access_token = ''
access_token_secret = ''

# Authentication with the Twitter API
auth = tweepy.OAuth1UserHandler(consumer_key, consumer_secret, access_token, access_token_secret)
api = tweepy.API(auth)

# Function for sentiment analysis
def analyze_sentiment(tweet):
    analysis = TextBlob(tweet)
    sentiment_score = analysis.sentiment.polarity
    print(f"Text: {tweet}\nSentiment Score: {sentiment_score}")
    return sentiment_score

# Function to save data to Excel file
def save_to_excel(data1, data2, sentiment_score):
    try:
        # Load existing workbook if it exists
        workbook = openpyxl.load_workbook("donnees1.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = "Donnée 1"
        sheet['B1'] = "Donnée 2"
        sheet['C1'] = "Score de sentiment"
    
    next_row = sheet.max_row + 1
    sheet[f'A{next_row}'] = data1
    sheet[f'B{next_row}'] = data2
    sheet[f'C{next_row}'] = sentiment_score
    
    workbook.save("donnees1.xlsx")
    print("Les données ont été ajoutées et enregistrées dans le fichier Excel.")

# Create a StreamingContext with a batch interval of 10 seconds
ssc = StreamingContext(sc, 10) 

# Create a DStream to represent the streaming data from a TCP socket
tweets = ssc.socketTextStream("localhost", 9000)  # RDD and DStream are used here

sentiment_scores = []

# Function to process each batch of data
def process_batch(batch_time, rdd):
    global sentiment_scores
    if not rdd.isEmpty():
        scores = rdd.collect()
        sentiment_scores.extend(scores)

# Call process_batch for each RDD in the DStream
tweets.foreachRDD(process_batch)

# Start the streaming context
ssc.start()

# Set up the GUI for the sentiment analysis
root = tk.Tk()
root.title("Twitter Sentiment Analysis")

# Function to stop the streaming context and close the GUI
def on_closing():
    ssc.stop()
    root.destroy()

# Function to analyze sentiment when the button is clicked
def perform_sentiment_analysis():
    text = entry.get()
    sentiment_score = analyze_sentiment(text)
    label.config(text=f"Sentiment Score: {sentiment_score:.2f}")
    save_to_excel(text, "", sentiment_score)  # Assuming only one input for simplicity

# Create an entry widget for user input
entry = tk.Entry(root, width=50)
entry.pack(pady=10)

# Create a button widget to trigger sentiment analysis
button = tk.Button(root, text="Analyze Sentiment", command=perform_sentiment_analysis)
button.pack(pady=5)

# Create a label widget to display sentiment score
label = tk.Label(root, text="")
label.pack(pady=10)

# Define behavior when the window is closed
root.protocol("WM_DELETE_WINDOW", on_closing)

# Run the GUI application
root.mainloop()
